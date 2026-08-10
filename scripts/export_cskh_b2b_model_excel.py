#!/usr/bin/env python3
"""Export RNOSAI CSKH B2B operating model to Excel (multi-sheet)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "exports" / "mo-hinh-cskh-b2b-rnosai.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="17692F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="17692F")
SUB_FONT = Font(italic=True, size=10, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_table(ws, start_row: int, headers: list[str], rows: list[list[str]]) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1
    return r


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def sheet_overview(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "1-Tong_quan"
    ws["A1"] = "MÔ HÌNH CSKH B2B — RNOSAI (PTT CRM)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Phiên bản 1.0 · 2026-08-04 · Nguồn: sales-b2b-lead-client-onboard-sop.md + ops-web lead-flow-kind"
    ws["A2"].font = SUB_FONT

    rows = [
        [
            "lead_flow_kind",
            "b2b_prospect",
            "Lead bán HĐ agency mới — DN prospect thuê PTT (Meta, SEO, Zalo…)",
        ],
        [
            "Khác với",
            "spa_operational",
            "Lead khách cuối từ Meta Ads của client đã active — SOP CSKH Spa 24h",
        ],
        [
            "Mục tiêu",
            "≤ 14 ngày",
            "Chốt HĐ → Customer → Agency Client active → sẵn sàng nhận lead vận hành",
        ],
        [
            "Staff URL",
            "rs.pttads.vn",
            "ops-web + Nest API (PostgreSQL source of truth)",
        ],
        [
            "Won status",
            "won",
            "KHÔNG dùng chot (chot = won lead spa end-user)",
        ],
        [
            "Review queue",
            "BR-CRM-003",
            "Deal > ngưỡng GDKD → bắt buộc review trước proposal",
        ],
    ]
    write_table(ws, 4, ["Khái niệm", "Giá trị", "Mô tả"], rows)
    autosize(ws)


def sheet_flow(wb: Workbook) -> None:
    ws = wb.create_sheet("2-Luong_nghiep_vu")
    ws["A1"] = "LUỒNG B2B SALES (6 bước trên UI + handoff)"
    ws["A1"].font = TITLE_FONT

    rows = [
        ["1", "Lead B2B", "moi", "/crm/leads/new · /crm/leads/[id]", "Referral / web / event / tạo tay · KHÔNG gắn agency_client_id client đang vận hành", "Sales / AM"],
        ["2", "B2 Liên hệ", "da_lien_he", "#funnel-b2 · LeadFunnelPanel", "Activity discovery · BANT sơ bộ · care pipeline B2 (≠ SLA spa 24h)", "Sales"],
        ["3", "Pre-sales", "dang_tu_van", "#funnel-presales · CRM-UC-005", "Record presales · stage: lead → consult → proposal · KH MKT sơ bộ", "Pre-sales"],
        ["4", "Intake BANT", "dang_tu_van", "/crm/intake?lead_id=", "Phiên discovery · catalog ngành/dịch vụ · gate BR-CRM-005", "Pre-sales / Sales"],
        ["5", "Proposal", "proposal", "/crm/proposals · CRM-UC-006", "Template + SKU · PDF/email · client accept · version audit", "Sales"],
        ["6", "Hợp đồng", "proposal", "#lead-contract · LeadContractPanel", "Draft → submit → approval → active · amount VND · service slug", "Sales · GDKD · Legal"],
        ["7", "Convert", "won", "/crm/customers · CRM-UC-007", "Promote HĐ → lead won · Customer record · lifecycle Onboard", "Sales / AM"],
        ["8", "Triển khai", "won", "/crm/service-delivery/[id]", "Service lifecycle · consult → onboard → deliver", "AM · Delivery"],
        ["9", "Agency Client", "active", "/agency/clients/new · SYS-UC-001", "Onboarding checklist 100% · channels Meta/Zalo · portal users", "AM · Tracking"],
        ["10", "Handoff CSKH Spa", "—", "SOP spa Meta 24h", "Client active → webhook Meta → lead vận hành (spa_operational)", "AM → CSKH rep"],
    ]
    write_table(
        ws,
        3,
        ["Bước", "Giai đoạn", "Status gợi ý", "Route / Module", "Mô tả", "Vai trò"],
        rows,
    )
    autosize(ws)


def sheet_status(wb: Workbook) -> None:
    ws = wb.create_sheet("3-Trang_thai_lead")
    rows = [
        ["moi", "Lead prospect mới", "Tạo lead · chưa qualify", "B2B + Spa"],
        ["da_lien_he", "Đã discovery / qualify", "Activity + BANT sơ bộ", "B2B + Spa"],
        ["dang_tu_van", "Đang pre-sales / tư vấn", "Intake · presales record", "B2B + Spa"],
        ["bao_gia", "Đang báo giá", "Chuẩn bị proposal", "B2B only"],
        ["dam_phan", "Đàm phán", "Negotiation trước ký", "B2B only"],
        ["proposal", "Đã / đang gửi báo giá", "Proposal sent · contract draft", "B2B only"],
        ["won", "HĐ ký — deal B2B", "Promote · Customer · onboard", "B2B only — KHÔNG chot"],
        ["chot", "Chốt spa end-user", "Won lead vận hành Meta", "Spa only — KHÔNG B2B"],
        ["lost", "Không chốt", "Audit note bắt buộc", "B2B + Spa"],
        ["pending_cleanup", "Dọn dẹp / duplicate", "Review queue release", "B2B + Spa"],
    ]
    write_table(ws, 1, ["Code status", "Ý nghĩa", "Khi nào dùng", "Phạm vi"], rows)
    autosize(ws)


def sheet_sla(wb: Workbook) -> None:
    ws = wb.create_sheet("4-Milestone_SLA")
    rows = [
        ["A", "Liên hệ + qualify", "≤ 2 ngày làm việc", "Activity + status da_lien_he", "Sales"],
        ["B", "Pre-sales / KH MKT sơ bộ", "≤ 5 ngày sau qualify", "crm_lead_presales + draft plan", "Pre-sales"],
        ["C", "Proposal gửi khách", "≤ 3 ngày sau pre-sales OK", "/crm/proposals · PDF version", "Sales"],
        ["D", "HĐ ký", "Theo deal", "LeadContractPanel · HĐ active", "Sales · Legal"],
        ["E", "Convert Customer", "≤ 1 ngày sau ký", "Lead won · /crm/customers", "Sales / AM"],
        ["F", "Client onboard active", "≤ 14 ngày sau ký", "/agency/clients · checklist 100%", "AM · Tracking"],
    ]
    write_table(ws, 1, ["Giai đoạn", "Mốc", "Hạn gợi ý", "Bằng chứng CRM", "Owner"], rows)
    autosize(ws)


def sheet_roles(wb: Workbook) -> None:
    ws = wb.create_sheet("5-Vai_tro")
    rows = [
        ["Sales / AM", "A → C", "Qualify · proposal · contract draft · F1–F2 agency", "crm_leads view/edit · proposals"],
        ["Pre-sales", "B", "Intake · presales advance · KH MKT sơ bộ", "intake · funnel presales"],
        ["GDKD", "Review · D3", "Review queue deal lớn · approval HĐ", "review-queue · contract approve"],
        ["Legal / Finance", "D3–D4", "Contract submit · signed_on", "leads-contract approvals"],
        ["Tracking / Tech", "F3 · F5", "Meta/Zalo token · campaign map · preflight", "agency channels · meta/tracking"],
        ["CSKH rep", "Handoff F8", "Lead vận hành spa sau client active", "cskh-board · spa SOP 24h"],
    ]
    write_table(ws, 1, ["Vai trò", "Giai đoạn", "Trách nhiệm", "Cap / module"], rows)
    autosize(ws)


def sheet_system(wb: Workbook) -> None:
    ws = wb.create_sheet("6-He_thong")
    rows = [
        ["ops-web", "3200", "Staff UI", "rs.pttads.vn"],
        ["ptt-crm-api", "3000", "Nest API", "/api/v1/* · /api/crm/*"],
        ["PostgreSQL", "—", "Source of truth", "crm_leads · crm_lead_activities · crm_lead_presales · contracts"],
        ["ptt_worker", "—", "Jobs nền", "alerts · digest · async tasks"],
        ["nginx", "443", "TLS + proxy", "/api/ → Nest · / → ops-web"],
    ]
    write_table(ws, 1, ["Thành phần", "Port", "Vai trò", "Ghi chú"], rows)

    ws.cell(row=9, column=1, value="API / Route chính (B2B)")
    ws.cell(row=9, column=1).font = Font(bold=True, size=12)

    api_rows = [
        ["GET", "/api/v1/leads/:id", "Chi tiết lead"],
        ["GET", "/api/v1/leads/:id/funnel", "Funnel snapshot · lead_flow_kind"],
        ["POST", "/api/v1/leads/:id/presales", "Tạo / advance pre-sales"],
        ["GET", "/api/crm/leads/:id/attribution", "Attribution ads (nếu có campaign)"],
        ["GET/POST", "/api/v1/leads/:id/contract", "HĐ trên lead"],
        ["GET", "/api/v1/leads/review-queue", "Inbox phải tra soát GDKD"],
        ["GET", "/crm/intake", "Intake BANT UI"],
        ["GET", "/crm/proposals", "Proposal UI"],
        ["GET", "/agency/clients/new", "Tạo agency client"],
    ]
    write_table(ws, 10, ["Method", "Path", "Mô tả"], api_rows)
    autosize(ws)


def sheet_compare(wb: Workbook) -> None:
    ws = wb.create_sheet("7-B2B_vs_Spa")
    rows = [
        ["lead_flow_kind", "b2b_prospect", "spa_operational"],
        ["Đối tượng", "DN prospect thuê agency", "Khách cuối (spa) từ Meta Lead Ads"],
        ["Won status", "won", "chot"],
        ["SLA chính", "Milestone 2–14 ngày (deal)", "15p / 4h / 24h (Meta 24h)"],
        ["Funnel UI", "LeadB2bSalesFlowBar (6 bước)", "LeadFunnelPanel B2 care + SLA panel"],
        ["Pre-sales", "Có (crm_lead_presales)", "Không"],
        ["Contract panel", "Có (LeadContractPanel)", "Không"],
        ["CSKH Enterprise board", "Không áp dụng", "Có (/crm/cskh-board)"],
        ["agency_client_id", "Null đến khi onboard", "Có thể có khi lead thuộc client active"],
        ["SOP", "sales-b2b-lead-client-onboard-sop.md", "cskh-spa-lead-meta-24h-sop.md"],
    ]
    write_table(ws, 1, ["Tiêu chí", "B2B Sales", "CSKH Spa Meta 24h"], rows)
    autosize(ws)


def sheet_data(wb: Workbook) -> None:
    ws = wb.create_sheet("8-Du_lieu_PG")
    rows = [
        ["crm_leads", "Lead master", "sqlite_lead_id · status · owner_id · meta_json · agency_client_id"],
        ["crm_lead_activities", "Activity / care log", "activity_type · care_status · care_stage_key"],
        ["crm_lead_presales", "Pre-sales record", "stage: lead|consult|proposal · service_slug"],
        ["crm_lead_presales_tasks", "Task pre-sales", "presales_id · stage · step_index"],
        ["crm_marketing_plans", "KH MKT sơ bộ", "presales_id · plan_kind draft"],
        ["crm_lead_contracts", "HĐ trên lead", "status draft|active · amount · service_slug"],
        ["crm_customers", "Customer sau won", "promote từ contract"],
        ["agency_clients", "Client vận hành", "onboarding checklist · channels · status active"],
    ]
    write_table(ws, 1, ["Bảng PG", "Mục đích", "Cột / quan hệ chính"], rows)

    ws.cell(row=12, column=1, value="Presales stages (LeadB2bSalesFlowBar)")
    ws.cell(row=12, column=1).font = Font(bold=True, size=12)
    stage_rows = [
        ["lead", "1", "Tạo presales record"],
        ["consult", "2", "Intake BANT OK · consult gate pass"],
        ["proposal", "3", "Sẵn sàng proposal / contract"],
    ]
    write_table(ws, 13, ["Stage", "Thứ tự", "Điều kiện"], stage_rows)
    autosize(ws)


def sheet_diagram(wb: Workbook) -> None:
    ws = wb.create_sheet("9-So_do_luong")
    ws["A1"] = "SƠ ĐỒ LUỒNG B2B (text)"
    ws["A1"].font = TITLE_FONT
    diagram = """
Lead B2B (referral/web/event/tạo tay)
    │
    ▼
[/crm/leads/[id]] Qualify · status: moi → da_lien_he → dang_tu_van
    │  LeadB2bSalesFlowBar: B2 Liên hệ
    ▼
[Pre-sales CRM-UC-005] crm_lead_presales · stage lead → consult → proposal
    │
    ▼
[/crm/intake] Intake BANT · catalog dịch vụ
    │
    ▼
[/crm/proposals CRM-UC-006] Proposal PDF · client accept · status proposal
    │
    ▼
[LeadContractPanel] HĐ draft → submit → approval → active
    │
    ▼
[CRM-UC-007] Promote → lead status WON · Customer · lifecycle Onboard
    │
    ▼
[/crm/service-delivery] Triển khai dịch vụ
    │
    ▼
[/agency/clients/new SYS-UC-001] Onboard · channels · checklist 100% → ACTIVE
    │
    ▼
Handoff → Lead vận hành spa Meta (spa_operational · SOP CSKH 24h)
"""
    ws["A3"] = diagram.strip()
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 90
    ws.row_dimensions[3].height = 420

    ws["A25"] = "Review queue (BR-CRM-003): deal lớn → funnel BLOCKED → GDKD /crm/leads/review-queue"
    ws["A25"].font = Font(bold=True, color="C0392B")


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_overview(wb)
    sheet_flow(wb)
    sheet_status(wb)
    sheet_sla(wb)
    sheet_roles(wb)
    sheet_system(wb)
    sheet_compare(wb)
    sheet_data(wb)
    sheet_diagram(wb)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
