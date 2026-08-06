#!/usr/bin/env python3
"""Export ma trận phân quyền CSKH-01 / KD-01 / MKT-01 → Excel + Markdown (in PDF).

Usage:
  python3 scripts/export_position_permission_matrix_kd_mkt_cskh.py
  python3 scripts/export_position_permission_matrix_kd_mkt_cskh.py --out-dir docs/exports
"""
from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from admin_page_permissions import (  # noqa: E402
    ADMIN_CRM_SECTIONS,
    _POSITION_DEFAULT,
    default_grants_for_position,
    grants_map_to_rows,
)
from cms_permissions import CMS_ACTIONS, CMS_ACTION_LABELS_VI  # noqa: E402
from ptt_ui_button_permissions import CRM_UI_BUTTON_BY_ID  # noqa: E402

POSITIONS: tuple[tuple[str, str, str], ...] = (
    ("CSKH-01", "Nhân viên CSKH vận hành", "CRM · CSKH vận hành 24h — lead spa/operational"),
    ("KD-01", "Account Manager (AM) B2B Sales", "CRM · B2B Sales — Intake → Giao Solution → Báo giá"),
    ("MKT-01", "Trưởng phòng Marketing / Solution", "CRM · Solution/MKT — Consult, R5, release Sales"),
)

EXTRA_ACTIONS: tuple[str, ...] = (
    "assign",
    "write",
    "settings",
    "compliance",
    "deliverability",
    "reports",
    "query",
    "commit",
    "run",
    "simulate",
)

EXTRA_LABELS: dict[str, str] = {
    "assign": "Gán / GDKD override",
    "write": "Ghi (email)",
    "settings": "Cài đặt (email)",
    "compliance": "Tuân thủ (email)",
    "deliverability": "Deliverability (email)",
    "reports": "Báo cáo (email)",
    "query": "Truy vấn AI",
    "commit": "Commit forecast",
    "run": "Chạy orchestrator",
    "simulate": "Mô phỏng workflow",
}

ALL_ACTIONS: tuple[str, ...] = CMS_ACTIONS + EXTRA_ACTIONS

EXTRA_SECTIONS: dict[str, dict[str, str]] = {
    "crm_email_mkt": {
        "label": "Email Marketing",
        "group": "CRM · Email",
        "page": "/email/hub",
        "description": "Campaign, journey, compliance — actions write/settings/…",
    },
}

P3_ROWS: tuple[tuple[str, str, str, str, str], ...] = (
    ("P3 Handoff", "crm_leads.edit", "Giao Solution/MKT →", "KD-01", "Stepper lead B2B sau Intake Go"),
    ("P3 Handoff", "crm_presales_solution.view", "Xem Solution queue / banner", "KD-01, MKT-01", "Theo dõi handoff"),
    ("P3 Handoff", "crm_presales_solution.edit", "Sửa Consult / L2 / R5", "MKT-01", "Workspace Tư vấn"),
    ("P3 Handoff", "crm_presales_solution.claim", "Nhận case", "MKT-01", "Queue hoặc stepper"),
    ("P3 Handoff", "crm_presales_solution.release", "Trả Sales — Báo giá →", "MKT-01", "Sau Consult+R5 ✓"),
    ("P3 Handoff", "crm_leads.assign", "GDKD override (review, handoff block)", "GDKD only", "Không mặc định KD-01"),
)


def _raw_grants(code: str) -> dict[str, frozenset[str]]:
    raw = _POSITION_DEFAULT.get(code, {})
    return {k: frozenset(v) for k, v in raw.items()}


def _expanded_grants(code: str) -> dict[str, list[str]]:
    """Merge CMS_ACTION-filtered grants with extra actions from raw defaults."""
    base = default_grants_for_position(code)
    raw = _raw_grants(code)
    out: dict[str, list[str]] = {k: list(v) for k, v in base.items()}
    for sid, acts in raw.items():
        if sid.startswith("crm_leads__"):
            continue
        merged = set(out.get(sid) or []) | {a for a in acts if a in EXTRA_ACTIONS}
        if merged:
            out[sid] = sorted(merged, key=lambda a: ALL_ACTIONS.index(a) if a in ALL_ACTIONS else 99)
    return out


def _action_cell(grants: dict[str, list[str]], section_id: str, action: str) -> str:
    if action in set(grants.get(section_id) or []):
        return "✓"
    btn = CRM_UI_BUTTON_BY_ID.get(section_id)
    if btn and action == btn["requires_action"] and action in set(grants.get(section_id) or []):
        return "✓"
    return ""


def _allowed_summary(raw: dict[str, frozenset[str]], section_id: str) -> str:
    acts = raw.get(section_id)
    if not acts:
        return "—"
    labels = [CMS_ACTION_LABELS_VI.get(a, EXTRA_LABELS.get(a, a)) for a in sorted(acts)]
    return ", ".join(labels)


def write_excel(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    yes_fill = PatternFill("solid", fgColor="D1FAE5")
    no_fill = PatternFill("solid", fgColor="F3F4F6")

    # —— Sheet Tổng quan ——
    ws0 = wb.active
    ws0.title = "Tổng quan"
    ws0.append(["Ma trận phân quyền PTT ops-web"])
    ws0.append(["Ngày xuất", str(date.today())])
    ws0.append(["Nguồn", "admin_page_permissions._POSITION_DEFAULT"])
    ws0.append([])
    ws0.append(["Mã", "Tên chức vụ", "Phạm vi"])
    for code, name, scope in POSITIONS:
        ws0.append([code, name, scope])
    ws0.append([])
    ws0.append(["Ghi chú"])
    ws0.append(["✓ = có quyền; ô trống = không"])
    ws0.append(["crm_leads.assign (GDKD) không nằm trong default KD-01/MKT-01/CSKH-01"])
    ws0.append(["P3 Solution: xem sheet «P3 Handoff»"])

    # —— Sheet P3 ——
    ws_p3 = wb.create_sheet("P3 Handoff")
    ws_p3.append(["Nhóm", "Cap (section.action)", "CTA / Menu", "Chức vụ", "Ghi chú"])
    for row in P3_ROWS:
        ws_p3.append(list(row))

    action_headers = [CMS_ACTION_LABELS_VI.get(a, a) for a in CMS_ACTIONS]
    extra_headers = [EXTRA_LABELS.get(a, a) for a in EXTRA_ACTIONS]

    for code, name, _scope in POSITIONS:
        ws = wb.create_sheet(code)
        grants = _expanded_grants(code)
        raw = _raw_grants(code)
        rows = grants_map_to_rows(grants)

        headers = [
            "Nhóm",
            "Loại",
            "Section / Nút",
            "Trang",
            "Mô tả",
            *action_headers,
            *extra_headers,
            "Tóm tắt quyền",
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row in rows:
            sid = str(row["section_id"])
            kind = str(row["row_kind"])
            if kind == "button_group":
                continue
            label = str(row["section_label"])
            group = str(row["group"])
            page = str(row.get("page") or "")
            desc = str(row.get("description") or "")
            parent_sid = str(row.get("parent_section") or sid)
            summary_sid = parent_sid if kind == "ui_button" else sid
            line = [group, kind, label, page, desc]
            for act in CMS_ACTIONS:
                val = "✓" if row["actions"].get(act) else ""
                line.append(val)
            for act in EXTRA_ACTIONS:
                val = "✓" if act in set(raw.get(summary_sid if kind != "ui_button" else sid) or []) else ""
                if kind == "ui_button" and act != "assign":
                    val = ""
                line.append(val)
            line.append(_allowed_summary(raw, summary_sid if kind == "section" else sid))
            ws.append(line)
            r = ws.max_row
            for c, act in enumerate(CMS_ACTIONS, start=6):
                if ws.cell(r, c).value == "✓":
                    ws.cell(r, c).fill = yes_fill
            for c, act in enumerate(EXTRA_ACTIONS, start=6 + len(CMS_ACTIONS)):
                if ws.cell(r, c).value == "✓":
                    ws.cell(r, c).fill = yes_fill

        for sid, meta in EXTRA_SECTIONS.items():
            acts = raw.get(sid)
            if not acts:
                continue
            line = [
                meta["group"],
                "section",
                meta["label"],
                meta["page"],
                meta.get("description", ""),
            ]
            for act in CMS_ACTIONS:
                line.append("✓" if act in acts else "")
            for act in EXTRA_ACTIONS:
                line.append("✓" if act in acts else "")
            line.append(_allowed_summary(raw, sid))
            ws.append(line)

        ws.freeze_panes = "A2"
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14 if col >= 6 else 22

    # —— Sheet so sánh ——
    ws_cmp = wb.create_sheet("So sánh")
    cmp_headers = ["Nhóm", "Section / Nút", "Trang"] + [p[0] for p in POSITIONS]
    ws_cmp.append(cmp_headers)
    for col in range(1, len(cmp_headers) + 1):
        cell = ws_cmp.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    grant_maps = {code: _expanded_grants(code) for code, _, _ in POSITIONS}
    raw_maps = {code: _raw_grants(code) for code, _, _ in POSITIONS}
    seen: set[str] = set()
    for sec in ADMIN_CRM_SECTIONS:
        sid = sec["id"]
        seen.add(sid)
        line = [sec["group"], sec["label"], sec.get("page", "")]
        for code, _, _ in POSITIONS:
            acts = raw_maps[code].get(sid) or frozenset()
            line.append(", ".join(sorted(acts)) if acts else "—")
        ws_cmp.append(line)
        for btn_id, btn in CRM_UI_BUTTON_BY_ID.items():
            if btn["parent_section"] != sid:
                continue
            bline = [sec["group"], f"  ↳ {btn['label']}", btn.get("page", "")]
            for code, _, _ in POSITIONS:
                g = grant_maps[code]
                req = btn["requires_action"]
                bline.append("✓" if req in set(g.get(btn_id) or []) else "—")
            ws_cmp.append(bline)

    ws_cmp.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_markdown(path: Path) -> None:
    lines: list[str] = []
    lines.append("# Ma trận phân quyền — CSKH-01 / KD-01 / MKT-01")
    lines.append("")
    lines.append(f"**Ngày xuất:** {date.today()}  ")
    lines.append("**Hệ thống:** PTT ops-web + Nest API (`section_id` × `action`)  ")
    lines.append("**Nguồn:** `admin_page_permissions._POSITION_DEFAULT`")
    lines.append("")
    lines.append("***")
    lines.append("")
    lines.append("## 1. Tổng quan chức vụ")
    lines.append("")
    lines.append("| Mã | Tên | Vai trò vận hành |")
    lines.append("|----|-----|------------------|")
    for code, name, scope in POSITIONS:
        lines.append(f"| **{code}** | {name} | {scope} |")
    lines.append("")
    lines.append("***")
    lines.append("")
    lines.append("## 2. P3 — Handoff Sales → Solution (bổ sung)")
    lines.append("")
    lines.append("| Cap | KD-01 (AM) | MKT-01 (Solution) | CSKH-01 |")
    lines.append("|-----|:----------:|:-----------------:|:-------:|")
    lines.append("| `crm_presales_solution.view` | ✓ | ✓ | — |")
    lines.append("| `crm_presales_solution.edit` | — | ✓ | — |")
    lines.append("| `crm_presales_solution.claim` | — | ✓ | — |")
    lines.append("| `crm_presales_solution.release` | — | ✓ | — |")
    lines.append("| Giao Solution (`crm_leads.edit` + stepper) | ✓ | — | — |")
    lines.append("| Consult read-only (AM khi handoff active) | ✓ | — | — |")
    lines.append("")
    lines.append("***")

    for code, name, scope in POSITIONS:
        raw = _raw_grants(code)
        lines.append(f"## 3.{POSITIONS.index((code, name, scope)) + 1} — {code}: {name}")
        lines.append("")
        lines.append(f"*{scope}*")
        lines.append("")
        current_group = ""
        for sec in ADMIN_CRM_SECTIONS:
            sid = sec["id"]
            acts = raw.get(sid)
            if not acts:
                continue
            if sec["group"] != current_group:
                current_group = sec["group"]
                lines.append(f"### {current_group}")
                lines.append("")
                lines.append("| Section | Trang | Quyền |")
                lines.append("|---------|-------|-------|")
            act_txt = ", ".join(
                CMS_ACTION_LABELS_VI.get(a, EXTRA_LABELS.get(a, a)) for a in sorted(acts)
            )
            lines.append(f"| {sec['label']} | `{sec.get('page', '')}` | {act_txt} |")
            btns = [b for b in CRM_UI_BUTTON_BY_ID.values() if b["parent_section"] == sid]
            for btn in btns:
                bid = btn["id"]
                btn_raw = raw.get(bid)
                if btn_raw is None:
                    grants = _expanded_grants(code)
                    allowed = btn["requires_action"] in set(grants.get(bid) or [])
                else:
                    allowed = bool(btn_raw)
                if allowed:
                    lines.append(
                        f"| ↳ **{btn['label']}** | `{btn.get('page', '')}` | ✓ ({btn['requires_action']}) |"
                    )
        for sid, meta in EXTRA_SECTIONS.items():
            acts = raw.get(sid)
            if not acts:
                continue
            if meta["group"] != current_group:
                current_group = meta["group"]
                lines.append(f"### {current_group}")
                lines.append("")
                lines.append("| Section | Trang | Quyền |")
                lines.append("|---------|-------|-------|")
            act_txt = ", ".join(
                CMS_ACTION_LABELS_VI.get(a, EXTRA_LABELS.get(a, a)) for a in sorted(acts)
            )
            lines.append(f"| {meta['label']} | `{meta['page']}` | {act_txt} |")
        lines.append("")
        lines.append("***")
        lines.append("")

    lines.append("## 4. Ký duyệt")
    lines.append("")
    lines.append("| Vai trò | Họ tên | Ngày | Chữ ký |")
    lines.append("|---------|--------|------|--------|")
    lines.append("| PO / Product Owner | | | |")
    lines.append("| GDKD Sales | | | |")
    lines.append("| Head Solution / MKT | | | |")
    lines.append("| IT / Admin hệ thống | | | |")
    lines.append("")
    lines.append("***")
    lines.append("")
    lines.append("*In file này ra PDF: mở bằng VS Code / browser → Print → Save as PDF.*")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export permission matrix Excel + Markdown")
    parser.add_argument(
        "--out-dir",
        default=str(ROOT / "docs" / "exports"),
        help="Output directory",
    )
    args = parser.parse_args()
    out_dir = Path(args.out_dir)
    stamp = date.today().isoformat()
    xlsx = out_dir / f"ma-tran-phan-quyen-CSKH-KD-MKT-{stamp}.xlsx"
    md = out_dir / f"ma-tran-phan-quyen-CSKH-KD-MKT-{stamp}.md"

    write_excel(xlsx)
    write_markdown(md)
    print(f"OK  Excel → {xlsx}")
    print(f"OK  Markdown (in PDF) → {md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
