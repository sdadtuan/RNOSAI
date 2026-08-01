#!/usr/bin/env python3
"""Generate RNOSAI BA spec workbook — quản lý màn hình, UC, test & traceability."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "samples" / "RNOSAI_BA_Spec.xlsx"
OUTPUT_FIXTURE = ROOT / "tests" / "fixtures" / "test_data" / "RNOSAI_BA_Spec.xlsx"

SCR_ID_RE = re.compile(r"SCR-[A-Z0-9]+-\d+")
UC_ID_RE = re.compile(r"[A-Z]+-UC-\d+")

from rnosai_ba_catalog_data import (  # noqa: E402
    BUSINESS_RULES,
    CODE_REGISTRY,
    MODULES,
    SCREENS,
    TEST_CASES,
    TODAY,
    TRACEABILITY,
    USE_CASES,
    VERSION,
    get_all_screen_details,
    get_all_use_case_details,
    manual_use_case_count,
)

# ── Theme ────────────────────────────────────────────────────────────────────
NAVY = "1F4E79"
NAVY_LIGHT = "D6E4F0"
ACCENT = "2E75B6"
WHITE = "FFFFFF"
ROW_ALT = "F5F9FC"
LINK_BLUE = "0563C1"
STATUS = {
    "Draft": "D9D9D9",
    "In progress": "FFF2CC",
    "Done": "C6EFCE",
    "Cancelled": "FFC7CE",
    "Pending": "E2EFDA",
    "Pass": "C6EFCE",
    "Fail": "FFC7CE",
}
PRIORITY = {"High": "FCE4D6", "Medium": "FFF2CC", "Low": "E2EFDA"}
MODULE_TAB = {
    "CRM": "548235",
    "META": "2E75B6",
    "ZALO": "0070C0",
    "SEO": "7030A0",
    "EM": "BF8F00",
    "PORTAL": "00B0F0",
    "AI": "C00000",
    "SVC": "375623",
    "AGENCY": "375623",
    "SYS": "44546A",
    "PLAT": "7F7F7F",
    "AUTH": "7F7F7F",
    "ADMIN": "44546A",
}
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

STATUS_LIST = '"Draft,In progress,Done,Cancelled"'
PRIORITY_LIST = '"High,Medium,Low"'
TEST_STATUS_LIST = '"Pending,Pass,Fail,Blocked,Skip"'
INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", underline=None) -> Font:
    return Font(bold=bold, size=size, color=color, underline=underline)


def _style_range(ws, row: int, col_start: int, col_end: int, fill=None, font=None, align=None) -> None:
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        cell.border = BORDER


def _auto_width(ws, min_w=10, max_w=48) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(c.value or "")) for c in col_cells[:120])
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def _table_header(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = _fill(NAVY)
        cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = CENTER
        cell.border = BORDER


def _section_title(ws, row: int, title: str, ncol: int = 6) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _fill(ACCENT)
    cell.font = _font(bold=True, color=WHITE, size=11)
    cell.alignment = Alignment(vertical="center")
    cell.border = BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_kv_block(ws, start_row: int, rows: list[tuple[str, str]], label_width: int = 22) -> int:
    _table_header(ws, start_row, ["Trường", "Nội dung"])
    r = start_row + 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = _font(bold=True)
        ws.cell(row=r, column=1).fill = _fill(NAVY_LIGHT)
        ws.cell(row=r, column=2, value=value)
        for col in (1, 2):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.alignment = WRAP
        ws.row_dimensions[r].height = max(18, min(60, 14 + len(str(value)) // 40 * 12))
        r += 1
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = 72
    return r


def _write_data_table(ws, start_row: int, headers: list[str], rows: list[list], zebra=True) -> int:
    _table_header(ws, start_row, headers)
    r = start_row + 1
    for i, row in enumerate(rows):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if zebra and i % 2 == 1:
                cell.fill = _fill(ROW_ALT)
        ws.row_dimensions[r].height = max(18, min(72, 14 + max(len(str(v)) for v in row) // 30 * 10))
        r += 1
    return r


def _apply_status_color(cell, value: str) -> None:
    key = str(value or "")
    if key in STATUS:
        cell.fill = _fill(STATUS[key])


def _apply_list_validation(ws, col: int, start_row: int, end_row: int, formula: str) -> None:
    if end_row < start_row:
        return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    for row in range(start_row, end_row + 1):
        dv.add(ws.cell(row=row, column=col))


def _setup_list_sheet(ws, headers: list[str], rows: list[list], status_col: int | None = None,
                      priority_col: int | None = None, freeze=True) -> None:
    _table_header(ws, 1, headers)
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if status_col and col == status_col:
                _apply_status_color(cell, str(val))
            if priority_col and col == priority_col and str(val) in PRIORITY:
                cell.fill = _fill(PRIORITY[str(val)])
        ws.row_dimensions[i].height = max(18, min(60, 16 + len(str(row[1] if len(row) > 1 else "")) // 35 * 10))
    if freeze:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    if status_col:
        _apply_list_validation(ws, status_col, 2, ws.max_row, STATUS_LIST)
    if priority_col:
        _apply_list_validation(ws, priority_col, 2, ws.max_row, PRIORITY_LIST)
    _auto_width(ws)


def _safe_sheet_name(raw: str, used: set[str]) -> str:
    """Excel sheet name: max 31 chars, no \\ / ? * [ ]"""
    name = INVALID_SHEET_CHARS.sub("-", raw.strip())[:31]
    if not name:
        name = "Sheet"
    base, n = name, 2
    while name in used:
        suffix = f"_{n}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(name)
    return name


def _entity_tab_color(entity_id: str) -> str:
    parts = entity_id.split("-")
    if len(parts) >= 2:
        key = parts[1] if parts[0] == "SCR" else parts[0]
        return MODULE_TAB.get(key, ACCENT)
    return ACCENT


def _parse_scr_ids(text: str) -> list[str]:
    return SCR_ID_RE.findall(str(text or ""))


def _parse_uc_ids(text: str) -> list[str]:
    return UC_ID_RE.findall(str(text or ""))


def _set_sheet_link(cell, sheet_name: str, display: str | None = None) -> None:
    """Internal hyperlink tới sheet trong cùng workbook."""
    label = display if display is not None else sheet_name
    cell.value = label
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=f"'{sheet_name}'!A1",
        display=str(label),
    )
    cell.font = _font(color=LINK_BLUE, underline="single")


def _set_entity_link(cell, entity_id: str, sheet_map: dict[str, str]) -> None:
    """Hyperlink từ mã SCR/UC tới sheet chi tiết (nếu có trong map)."""
    sheet_name = sheet_map.get(entity_id)
    if sheet_name:
        _set_sheet_link(cell, sheet_name, entity_id)
    else:
        cell.value = entity_id


def _link_id_cell(
    cell,
    text: str,
    sheet_map: dict[str, str],
    *,
    parse_ids,
) -> None:
    """Một ID → hyperlink; nhiều ID → giữ text (mở từ sheet chi tiết)."""
    raw = str(text or "")
    ids = parse_ids(raw)
    if len(ids) == 1 and ids[0] in sheet_map:
        _set_entity_link(cell, ids[0], sheet_map)
    else:
        cell.value = raw


def _write_linked_entities_table(
    ws,
    start_row: int,
    title: str,
    entity_ids: list[str],
    sheet_map: dict[str, str],
) -> int:
    """Bảng entity → link sheet (dùng trên sheet chi tiết)."""
    if not entity_ids:
        return start_row
    ws.cell(row=start_row, column=1, value=title).font = _font(bold=True, size=11)
    start_row += 1
    start_row = _write_data_table(
        ws,
        start_row,
        ["Mã", "Mở sheet spec"],
        [[eid, sheet_map.get(eid, "—")] for eid in entity_ids],
    )
    link_row = start_row - len(entity_ids)
    for offset, eid in enumerate(entity_ids):
        r = link_row + offset
        if eid in sheet_map:
            _set_entity_link(ws.cell(row=r, column=1), eid, sheet_map)
            _set_sheet_link(ws.cell(row=r, column=2), sheet_map[eid], f"→ {sheet_map[eid]}")
    return start_row + 1


def _write_detail_nav(
    ws,
    *,
    back_catalog: str,
    related_uc_ids: list[str] | None = None,
    related_scr_ids: list[str] | None = None,
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> int:
    """Khối điều hướng đầu sheet chi tiết."""
    row = 1
    ws.cell(row=row, column=1, value="← Catalog:").font = _font(bold=True, size=10)
    _set_sheet_link(ws.cell(row=row, column=2), back_catalog, back_catalog)
    row += 2
    if related_uc_ids:
        row = _write_linked_entities_table(
            ws, row, "Use case liên quan (click mở sheet UC)", related_uc_ids, uc_sheet_map
        )
    if related_scr_ids:
        row = _write_linked_entities_table(
            ws, row, "Màn hình liên quan (click mở sheet SCR)", related_scr_ids, scr_sheet_map
        )
    return row


def _write_screen_detail_sheet(
    ws,
    scr_id: str,
    detail: dict,
    br_map: dict[str, str],
    *,
    linked_uc_ids: list[str],
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> None:
    ws.sheet_properties.tabColor = _entity_tab_color(scr_id)
    row = _write_detail_nav(
        ws,
        back_catalog="01_DanhSach_ManHinh",
        related_uc_ids=linked_uc_ids,
        scr_sheet_map=scr_sheet_map,
        uc_sheet_map=uc_sheet_map,
    )
    badge = (
        "Spec deep (page.tsx)"
        if detail.get("_deep")
        else ("Spec thủ công P0" if not detail.get("_auto") else "Spec auto")
    )
    row = _section_title(ws, row, f"{scr_id} — {detail['meta'][1][1]}", 5)
    ws.cell(row=row, column=1, value=f"Loại spec: {badge}").font = _font(bold=True, size=10, color="555555")
    row += 2
    row = _write_kv_block(ws, row, detail["meta"]) + 1

    ws.cell(row=row, column=1, value="Thành phần UI").font = _font(bold=True, size=11)
    row += 1
    row = _write_data_table(
        ws, row,
        ["STT", "Thành phần", "Loại", "Bắt buộc", "Mô tả / Hành vi"],
        detail["ui"],
    ) + 1

    rule_rows = [[rid, br_map.get(rid, "")] for rid in detail["rules"]]
    ws.cell(row=row, column=1, value="Quy tắc nghiệp vụ liên quan").font = _font(bold=True, size=11)
    row += 1
    _write_data_table(ws, row, ["Mã rule", "Mô tả"], rule_rows)
    _auto_width(ws, max_w=36)


def _write_uc_detail_sheet(
    ws,
    uc_id: str,
    detail: dict,
    br_map: dict[str, str],
    *,
    linked_scr_ids: list[str],
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> None:
    ws.sheet_properties.tabColor = _entity_tab_color(uc_id)
    row = _write_detail_nav(
        ws,
        back_catalog="03_DanhSach_UseCase",
        related_scr_ids=linked_scr_ids,
        scr_sheet_map=scr_sheet_map,
        uc_sheet_map=uc_sheet_map,
    )
    badge = "Spec thủ công" if detail.get("_manual") else "Spec auto"
    row = _section_title(ws, row, f"{uc_id} — {detail['meta'][1][1]}", 4)
    ws.cell(row=row, column=1, value=f"Loại spec: {badge}").font = _font(bold=True, size=10, color="555555")
    row += 2
    row = _write_kv_block(ws, row, detail["meta"]) + 1

    ws.cell(row=row, column=1, value="Luồng chính").font = _font(bold=True, size=11)
    row += 1
    row = _write_data_table(ws, row, ["Bước", "Mô tả"], detail["main_flow"]) + 1

    ws.cell(row=row, column=1, value="Luồng thay thế / ngoại lệ").font = _font(bold=True, size=11)
    row += 1
    row = _write_data_table(ws, row, ["Mã", "Mô tả"], detail["alt_flow"]) + 1

    ws.cell(row=row, column=1, value="Dữ liệu vào / ra").font = _font(bold=True, size=11)
    row += 1
    row = _write_data_table(ws, row, ["Loại", "Nội dung"], detail["io"]) + 1

    rule_rows = [[rid, br_map.get(rid, "")] for rid in detail["rules"]]
    ws.cell(row=row, column=1, value="Quy tắc nghiệp vụ").font = _font(bold=True, size=11)
    row += 1
    _write_data_table(ws, row, ["Mã rule", "Mô tả"], rule_rows)
    _auto_width(ws, max_w=36)


def _build_tong_quan(wb: Workbook, scr_sheet_map: dict[str, str], uc_sheet_map: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "00_TongQuan"
    ws.sheet_properties.tabColor = NAVY

    ws.merge_cells("A1:F1")
    ws["A1"] = "RNOSAI — Bộ tài liệu BA & Quản lý dự án"
    ws["A1"].font = _font(bold=True, size=18, color=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Revenue Operating System + AI · Phiên bản {VERSION} · Xuất {TODAY}"
    ws["A2"].font = _font(size=11, color="555555")

    info = [
        ("Dự án", "RNOSAI (Revenue Operating System + AI)"),
        ("Repository", "https://github.com/sdadtuan/RNOSAI"),
        ("Spec master", "docs/specs/RNOSAI-BA-Master-Spec.md"),
        ("Module annexes", "docs/specs/modules/RNOSAI-BA-*-UseCases.md"),
        ("Ops-web routes", "services/ops-web/src/app/crm/*"),
        ("API backend", "services/ptt-crm-api"),
        ("Mục đích file", "Quản lý màn hình · UC · BR · test · traceability"),
        ("Cấu trúc chi tiết", f"Mỗi SCR = 1 sheet · Mỗi UC = 1 sheet ({len(scr_sheet_map)} + {len(uc_sheet_map)} sheets)"),
        ("Điều hướng", "Click mã SCR/UC hoặc cột «Sheet spec» → mở sheet chi tiết"),
        ("Quy ước mã", "SCR / UC / TC / BR / API / RNOS / P0"),
    ]
    r = 4
    for label, val in info:
        ws.cell(row=r, column=1, value=label).font = _font(bold=True)
        ws.cell(row=r, column=1).fill = _fill(NAVY_LIGHT)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Thống kê nội dung").font = _font(bold=True, size=12)
    r += 1
    stats = [
        ("Màn hình (SCR)", len(SCREENS)),
        ("Sheet chi tiết SCR", len(scr_sheet_map)),
        ("Use case (UC)", len(USE_CASES)),
        ("Sheet chi tiết UC", len(uc_sheet_map)),
        ("Business rule (BR)", len(BUSINESS_RULES)),
        ("Test case (TC)", len(TEST_CASES)),
        ("Liên kết traceability", len(TRACEABILITY)),
        ("Tổng sheet workbook", 7 + len(scr_sheet_map) + len(uc_sheet_map)),
    ]
    _table_header(ws, r, ["Hạng mục", "Số lượng"])
    r += 1
    for name, cnt in stats:
        ws.cell(row=r, column=1, value=name).border = BORDER
        ws.cell(row=r, column=2, value=cnt).border = BORDER
        ws.cell(row=r, column=2).alignment = CENTER
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Danh sách sheet catalog & matrix").font = _font(bold=True, size=12)
    r += 1
    sheets = [
        ("00_DanhSach_Ma", "Quy ước mã chuẩn"),
        ("01_DanhSach_ManHinh", "Catalog màn hình + link sheet SCR"),
        ("02_Index_ManHinh", "Index nhanh → sheet chi tiết SCR"),
        ("SCR-*", f"{len(scr_sheet_map)} sheet — mỗi màn hình một sheet"),
        ("03_DanhSach_UseCase", "Catalog use case + link sheet UC"),
        ("04_Index_UseCase", "Index nhanh → sheet chi tiết UC"),
        ("*-UC-*", f"{len(uc_sheet_map)} sheet — mỗi use case một sheet"),
        ("05_Traceability_Matrix", "BR ↔ SCR ↔ UC ↔ TC"),
        ("06_TestCase", "Kịch bản kiểm thử UAT/E2E"),
        ("07_Change_Log", "Lịch sử thay đổi tài liệu"),
    ]
    _table_header(ws, r, ["Sheet / Nhóm", "Mô tả"])
    r += 1
    for sheet, desc in sheets:
        c1 = ws.cell(row=r, column=1, value=sheet)
        c1.border = BORDER
        if not sheet.endswith("*"):
            _set_sheet_link(c1, sheet, sheet)
        ws.cell(row=r, column=2, value=desc).border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 58


def _build_00_danh_sach_ma(wb: Workbook) -> None:
    ws = wb.create_sheet("00_DanhSach_Ma")
    ws.sheet_properties.tabColor = ACCENT
    headers = ["Loại", "Tiền tố", "Ví dụ", "Mô tả / Phạm vi"]
    rows = [[a, b, c, d] for a, b, c, d in CODE_REGISTRY]
    _setup_list_sheet(ws, headers, rows)
    ws.column_dimensions["D"].width = 42


def _build_01_danh_sach_man_hinh(
    wb: Workbook,
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> None:
    ws = wb.create_sheet("01_DanhSach_ManHinh")
    ws.sheet_properties.tabColor = "548235"
    headers = [
        "Mã màn hình", "Tên màn hình", "Module", "Route", "Vai trò sử dụng",
        "Trạng thái", "Use case liên quan", "Version", "Owner", "Priority",
        "Parity / RNOS", "Last Updated", "Ghi chú triển khai", "→ Sheet spec",
    ]
    rows = [list(row) + [scr_sheet_map[str(row[0])]] for row in SCREENS]
    _table_header(ws, 1, headers)
    for i, row in enumerate(rows, start=2):
        scr_id = str(row[0])
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if col == 1:
                _set_entity_link(cell, scr_id, scr_sheet_map)
            if col == 7:
                _link_id_cell(cell, str(val), uc_sheet_map, parse_ids=_parse_uc_ids)
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 6:
                _apply_status_color(cell, str(val))
            if col == 10 and str(val) in PRIORITY:
                cell.fill = _fill(PRIORITY[str(val)])
            if col == 14:
                _set_sheet_link(cell, str(val), f"→ {val}")
        ws.row_dimensions[i].height = max(18, min(60, 16 + len(str(row[1])) // 35 * 10))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _apply_list_validation(ws, 6, 2, ws.max_row, STATUS_LIST)
    _apply_list_validation(ws, 10, 2, ws.max_row, PRIORITY_LIST)
    _auto_width(ws)


def _build_02_index_man_hinh(wb: Workbook, scr_sheet_map: dict[str, str]) -> None:
    ws = wb.create_sheet("02_Index_ManHinh")
    ws.sheet_properties.tabColor = "548235"
    screen_by_id = {str(r[0]): r for r in SCREENS}
    headers = ["Mã SCR", "Tên màn hình", "Module", "Route", "Trạng thái", "→ Sheet spec"]
    rows = []
    for scr_id in sorted(scr_sheet_map):
        row = screen_by_id[scr_id]
        rows.append([scr_id, row[1], row[2], row[3], row[5], scr_sheet_map[scr_id]])
    _table_header(ws, 1, headers)
    for i, row in enumerate(rows, start=2):
        scr_id = str(row[0])
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if col == 1:
                _set_entity_link(cell, scr_id, scr_sheet_map)
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 5:
                _apply_status_color(cell, str(val))
            if col == 6:
                _set_sheet_link(cell, str(val), f"→ {val}")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_width(ws)


def _build_scr_detail_sheets(
    wb: Workbook,
    br_map: dict[str, str],
    used: set[str],
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> dict[str, str]:
    sheet_map: dict[str, str] = {}
    screen_by_id = {str(r[0]): r for r in SCREENS}
    for scr_id, detail in sorted(get_all_screen_details().items()):
        sheet_name = scr_sheet_map[scr_id]
        ws = wb.create_sheet(sheet_name)
        used.add(sheet_name)
        linked_ucs = _parse_uc_ids(str(screen_by_id.get(scr_id, ["", ""] * 7)[6]))
        _write_screen_detail_sheet(
            ws,
            scr_id,
            detail,
            br_map,
            linked_uc_ids=linked_ucs,
            scr_sheet_map=scr_sheet_map,
            uc_sheet_map=uc_sheet_map,
        )
        sheet_map[scr_id] = sheet_name
    return sheet_map


def _build_03_danh_sach_use_case(
    wb: Workbook,
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> None:
    ws = wb.create_sheet("03_DanhSach_UseCase")
    ws.sheet_properties.tabColor = "BF8F00"
    headers = [
        "Mã UC", "Tên use case", "Màn hình", "Actor chính", "Priority",
        "Trạng thái", "Pre-condition", "Post-condition", "Business Rule",
        "Owner", "Sprint/Wave", "Trace ref", "→ Sheet spec",
    ]
    rows = [list(row) + [uc_sheet_map[str(row[0])]] for row in USE_CASES]
    _table_header(ws, 1, headers)
    for i, row in enumerate(rows, start=2):
        uc_id = str(row[0])
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if col == 1:
                _set_entity_link(cell, uc_id, uc_sheet_map)
            if col == 3:
                _link_id_cell(cell, str(val), scr_sheet_map, parse_ids=_parse_scr_ids)
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 6:
                _apply_status_color(cell, str(val))
            if col == 5 and str(val) in PRIORITY:
                cell.fill = _fill(PRIORITY[str(val)])
            if col == 13:
                _set_sheet_link(cell, str(val), f"→ {val}")
        ws.row_dimensions[i].height = max(18, min(60, 16 + len(str(row[1])) // 35 * 10))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _apply_list_validation(ws, 6, 2, ws.max_row, STATUS_LIST)
    _apply_list_validation(ws, 5, 2, ws.max_row, PRIORITY_LIST)
    _auto_width(ws)


def _build_04_index_use_case(wb: Workbook, uc_sheet_map: dict[str, str]) -> None:
    ws = wb.create_sheet("04_Index_UseCase")
    ws.sheet_properties.tabColor = "BF8F00"
    uc_by_id = {str(r[0]): r for r in USE_CASES}
    headers = ["Mã UC", "Tên use case", "Module prefix", "Priority", "Trạng thái", "→ Sheet spec"]
    rows = []
    for uc_id in sorted(uc_sheet_map):
        row = uc_by_id[uc_id]
        prefix = uc_id.split("-UC-")[0]
        rows.append([uc_id, row[1], prefix, row[4], row[5], uc_sheet_map[uc_id]])
    _table_header(ws, 1, headers)
    for i, row in enumerate(rows, start=2):
        uc_id = str(row[0])
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if col == 1:
                _set_entity_link(cell, uc_id, uc_sheet_map)
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 5:
                _apply_status_color(cell, str(val))
            if col == 6:
                _set_sheet_link(cell, str(val), f"→ {val}")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _auto_width(ws)


def _build_uc_detail_sheets(
    wb: Workbook,
    br_map: dict[str, str],
    used: set[str],
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> dict[str, str]:
    sheet_map: dict[str, str] = {}
    uc_by_id = {str(r[0]): r for r in USE_CASES}
    for uc_id, detail in sorted(get_all_use_case_details().items()):
        sheet_name = uc_sheet_map[uc_id]
        ws = wb.create_sheet(sheet_name)
        used.add(sheet_name)
        linked_scrs = _parse_scr_ids(str(uc_by_id.get(uc_id, [""] * 3)[2]))
        _write_uc_detail_sheet(
            ws,
            uc_id,
            detail,
            br_map,
            linked_scr_ids=linked_scrs,
            scr_sheet_map=scr_sheet_map,
            uc_sheet_map=uc_sheet_map,
        )
        sheet_map[uc_id] = sheet_name
    return sheet_map


def _build_05_traceability(
    wb: Workbook,
    scr_sheet_map: dict[str, str],
    uc_sheet_map: dict[str, str],
) -> None:
    ws = wb.create_sheet("05_Traceability_Matrix")
    ws.sheet_properties.tabColor = "7030A0"
    headers = ["BR ID", "Màn hình (SCR)", "Use case (UC)", "Test case (TC)", "Trạng thái coverage"]
    _table_header(ws, 1, headers)
    for i, row in enumerate(TRACEABILITY, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 2:
                _link_id_cell(cell, str(val), scr_sheet_map, parse_ids=_parse_scr_ids)
            if col == 3:
                _link_id_cell(cell, str(val), uc_sheet_map, parse_ids=_parse_uc_ids)
            if col == 5:
                _apply_status_color(cell, str(val))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _apply_list_validation(ws, 5, 2, ws.max_row, STATUS_LIST)
    _auto_width(ws)


def _build_06_test_case(wb: Workbook, uc_sheet_map: dict[str, str]) -> None:
    ws = wb.create_sheet("06_TestCase")
    ws.sheet_properties.tabColor = "C00000"
    headers = [
        "Mã TC", "Mã UC", "Tên test", "Bước test", "Kết quả mong đợi",
        "Kết quả thực tế", "Trạng thái", "Priority", "Fixture / Evidence",
    ]
    _table_header(ws, 1, headers)
    for i, row in enumerate(TEST_CASES, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            cell.alignment = WRAP
            if col == 2:
                _link_id_cell(cell, str(val), uc_sheet_map, parse_ids=_parse_uc_ids)
            if i % 2 == 0:
                cell.fill = _fill(ROW_ALT)
            if col == 7:
                _apply_status_color(cell, str(val))
            if col == 8 and str(val) == "P0":
                cell.fill = _fill("FCE4D6")
        ws.row_dimensions[i].height = max(36, min(80, 20 + len(str(row[3])) // 25 * 12))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _apply_list_validation(ws, 7, 2, ws.max_row, TEST_STATUS_LIST)
    _auto_width(ws)
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["I"].width = 32


def _build_07_change_log(wb: Workbook) -> None:
    ws = wb.create_sheet("07_Change_Log")
    ws.sheet_properties.tabColor = "7F7F7F"
    rows = [
        [TODAY, VERSION, f"v2.2: Deep-spec Portal — 17 SCR portal-web ({len(SCREENS)} SCR total)", "BA / AI Agent", "rnosai_ba_scr_details_portal.py"],
        ["2026-07-28", "2.1", "Deep-spec EM module — 20 SCR", "BA / AI Agent", "rnosai_ba_scr_details_em.py"],
        ["2026-07-28", "1.7", f"Batch P1 +44 SCR +5 PORTAL UC", "BA / AI Agent", "rnosai_ba_catalog_data.py"],
        ["2026-07-27", "1.6", "Mỗi SCR/UC = 1 sheet; index + hyperlink", "BA / AI Agent", "build_ba_spec_workbook.py"],
        ["2026-07-27", "1.5", f"100% manual UC annexes — {manual_use_case_count()}/{len(USE_CASES)} UC", "BA", "docs/specs/modules/"],
        ["2026-07-27", "1.1", f"Catalog đầy đủ: {len(SCREENS)} SCR, {len(USE_CASES)} UC", "BA", "Auto-generated UC blocks"],
    ]
    _setup_list_sheet(ws, ["Ngày", "Phiên bản", "Nội dung thay đổi", "Người sửa", "Ghi chú"], rows)


def build_ba_spec_workbook() -> Workbook:
    wb = Workbook()
    br_map = {b[0]: b[1] for b in BUSINESS_RULES}

    reserved: set[str] = set()
    scr_sheet_map = {
        scr_id: _safe_sheet_name(scr_id, reserved) for scr_id in sorted(get_all_screen_details())
    }
    uc_sheet_map = {
        uc_id: _safe_sheet_name(uc_id, reserved) for uc_id in sorted(get_all_use_case_details())
    }

    _build_tong_quan(wb, scr_sheet_map, uc_sheet_map)
    _build_00_danh_sach_ma(wb)
    _build_01_danh_sach_man_hinh(wb, scr_sheet_map, uc_sheet_map)
    _build_02_index_man_hinh(wb, scr_sheet_map)

    used_names = set(wb.sheetnames) | reserved
    actual_scr_map = _build_scr_detail_sheets(
        wb, br_map, used_names, scr_sheet_map, uc_sheet_map
    )

    _build_03_danh_sach_use_case(wb, scr_sheet_map, uc_sheet_map)
    _build_04_index_use_case(wb, uc_sheet_map)

    used_names = set(wb.sheetnames)
    actual_uc_map = _build_uc_detail_sheets(
        wb, br_map, used_names, scr_sheet_map, uc_sheet_map
    )

    _build_05_traceability(wb, scr_sheet_map, uc_sheet_map)
    _build_06_test_case(wb, uc_sheet_map)
    _build_07_change_log(wb)

    assert actual_scr_map == scr_sheet_map
    assert actual_uc_map == uc_sheet_map
    return wb


def main() -> None:
    wb = build_ba_spec_workbook()
    for path in (OUTPUT, OUTPUT_FIXTURE):
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        print(f"Wrote {path} ({len(wb.sheetnames)} sheets, v{VERSION})")


if __name__ == "__main__":
    main()
